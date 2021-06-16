from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

import openpyxl

path = "./White box Testing_new16062021.xlsx"

lab2 = openpyxl.load_workbook(path)
register_test = lab2["Function1"]
columns = register_test.max_column
rows = register_test.max_row
print(rows)

exception_case = [rows-5, rows-6]

# Tạo ra danh sách các test case
test_case = []
for c in range(6, columns):
    unit = []
    for r in range(13, rows-4):
        checked = register_test.cell(row=r, column=c).value
        if checked == "O":
            # print("row number ", r)
            unit.append(r)

    if len(unit) == 0:
        pass
    else:
        test_case.append(unit)

# Config Driver
driver = webdriver.Chrome(executable_path=r"D:\UNIVERSITY\Nam_3\Nam3_2\SE113\chromedriver_win32\chromedriver.exe")
driver.maximize_window()
driver.get("http://127.0.0.1:8000/register/")

# username = driver.find_element_by_id("id_username")
# username.clear()
# username.send_keys("abc_test")

# first_name = driver.find_element_by_id("id_first_name")
# first_name.clear()
# first_name.send_keys("abc")

# last_name = driver.find_element_by_id("id_last_name")
# last_name.clear()
# last_name.send_keys("nguyen")

# email = driver.find_element_by_id("id_email")
# email.clear()
# email.send_keys("abc@gmail.com")

# psw_1 = driver.find_element_by_id("id_password1")
# psw_1.clear()
# psw_1.send_keys("")

# psw_2 = driver.find_element_by_id("id_password2")
# psw_2.clear()
# psw_2.send_keys("pipinstall")


# submit_btn = driver.find_element_by_class_name("submitButton")
# submit_btn.click()

# myValidationmsg = driver.find_element(By.NAME, 'password1').get_attribute("validationMessage")
# print(myValidationmsg)

# driver.close()

def sendKeys(value):
    if value:
        return value
    else:
        return ""
    


# Automation Test Cases
i = 0
untested_cases = 0
failed = 0
succeed = 0
for test in test_case:
    i += 1
    print("Testing:", i)
    # if test[7] in exception_case:
    #     untested_cases += 1
    #     continue
    # else:
    username = driver.find_element_by_id("id_username")
    username.clear()
    username.send_keys(sendKeys(register_test.cell(row=test[3], column=4).value))

    first_name = driver.find_element_by_id("id_first_name")
    first_name.clear()
    first_name.send_keys(sendKeys(register_test.cell(row=test[0], column=4).value))

    last_name = driver.find_element_by_id("id_last_name")
    last_name.clear()
    last_name.send_keys(sendKeys(register_test.cell(row=test[1], column=4).value))

    email = driver.find_element_by_id("id_email")
    email.clear()
    email.send_keys(sendKeys(register_test.cell(row=test[2], column=4).value))

    psw_1 = driver.find_element_by_id("id_password1")
    psw_1.clear()
    psw_1.send_keys(sendKeys(register_test.cell(row=test[4], column=4).value))

    psw_2 = driver.find_element_by_id("id_password2")
    psw_2.clear()
    psw_2.send_keys(sendKeys(register_test.cell(row=test[5], column=4).value))

    submit_btn = driver.find_element_by_class_name("submitButton")
    submit_btn.click()
    if test[7] not in exception_case:
        alert_msg = driver.switch_to.alert
        WebDriverWait(driver, 10).until(EC.alert_is_present())

        if alert_msg.text == register_test.cell(row=test[7], column=4).value:
            print("Test case", i, "succeed!")
            succeed += 1
        else:
            print("Test case", i, "failed!")
            failed += 1
        print("Actual alert: " + alert_msg.text)
        alert_msg.accept()
    elif test[7] == rows - 5:
        validation_msg = driver.find_element(By.NAME, 'email').get_attribute("validationMessage")
        if validation_msg == register_test.cell(row=test[7], column=4).value:
            print("Test case", i, "succeed!")
            succeed += 1
        else:
            print("Test case", i, "failed!")
            failed += 1
        print("Actual Message: " + validation_msg)
        continue
    else:
        cmpMsg = register_test.cell(row=test[7], column=4).value
        validMsg_psw1 = driver.find_element(By.NAME, 'password1').get_attribute("validationMessage")
        validMsg_psw2 = driver.find_element(By.NAME, 'password2').get_attribute("validationMessage")
        validMsg_username = driver.find_element(By.NAME, 'username').get_attribute("validationMessage")
        if (validMsg_psw1 == cmpMsg or validMsg_psw2 == cmpMsg or validMsg_username):
            print("Test case", i, "succeed!")
            succeed += 1
        else:
            print("Test case", i, "failed!")
            failed += 1
        print("Actual Message: " + validMsg_psw1 or validMsg_psw2 or validMsg_username) 
        continue


driver.close()

print("--------------------------")
print("Report: Register Testing")
print("Tested cases:", len(test_case) - untested_cases)
print("Untested cases:", untested_cases)
print("Failed tests:", failed)
print("Succeed tests:", succeed)