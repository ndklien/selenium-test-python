from selenium import webdriver
from selenium.webdriver.common.keys import Keys

import openpyxl
import time

path = "./White box Testing.xlsx"

lab2 = openpyxl.load_workbook(path)

search = lab2["Function5"]

columns = search.max_column
print(columns)

test_cases = []
for c in range(6, 9):
    case = []
    for r in range(13, 16):
        checked = search.cell(row=r, column=c).value
        if checked == "O":
            case.append(r)
    test_cases.append(case)
# print(test_cases)

# Config WebDriver
driver = webdriver.Chrome(executable_path=r"D:\UNIVERSITY\Nam_3\Nam3_2\SE113\chromedriver_win32\chromedriver.exe")
driver.maximize_window()
driver.get("http://127.0.0.1:8000/")

# results = driver.find_element_by_class_name("row")
i = 1
for case in test_cases:
    querySearch = driver.find_element_by_name("q")
    querySearch.clear()
    query = search.cell(row=case[0], column=4).value
    querySearch.send_keys(query)
    querySearch.send_keys(Keys.RETURN)
    result = driver.find_element_by_xpath("//div[@class='row']")
    if result:
        print("Test case",  i, "Passed")
    i += 1
driver.close()
